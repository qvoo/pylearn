from openpyxl import Workbook

#创建一个新的工作簿
wb = Workbook()

#取出原来默认的工作表   创建多个工作表
wb.remove(wb.active)

ws1=wb.create_sheet("第一个表")
ws2=wb.create_sheet("第二个表")

#向每个工作表中写入数据
ws1.append(["姓名","年龄","性别"])
ws1.append(["张三",20,"男"])

ws2.append(["学校","书院","班级"])
ws2.append(["海南大学","软件工程学院","2023"])


#保存工作簿到文件
wb.save("多表数据.xlsx")